"""
Task 02 — Pasif Honeytoken Enjeksiyon Modülü Testleri (TDD).

Test edilen bileşen: mirage.honeytoken
Amaç: Bir DataFrame'i (veya sentetik veriyi) XLSX dosyasına çevirirken,
dosya XML yapısına PASSIF bir tracking URL'i gömmek. Dosya açıldığında
Excel bu URL'e zararsız bir HTTP GET isteği atar.

KRİTİK GÜVENLİK İNVARİANTLARI (her testte doğrulanır):
  1. Hiçbir PowerShell/DDE/macro/DNS-tunneling payload'ı yoktur.
  2. Dosya açıldığında makinede HİÇBİR kod çalışmaz — yalnızca HTTP GET.
  3. URL benzersiz bir token içerir (her üretimde farklı).
  4. Dosya yapısı geçerli xlsx olarak kalır (Pandas/openpyxl sorunsuz okur).

Çalıştırma:
    python3 -m pytest test_honeytoken.py -v
    # veya pytest yoksa:
    python3 test_honeytoken.py
"""
from __future__ import annotations

import io
import re
import sys
import unittest
import uuid
import zipfile
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from mirage.honeytoken import (
    inject_honeytoken,
    HoneytokenRegistry,
    HoneytokenRecord,
    MIRAGE_FORBIDDEN_PATTERNS,
)


# --- Test verisi -------------------------------------------------------------
def make_sample_df(n: int = 50) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "user_id": [f"usr-{i:04d}" for i in range(n)],
            "age": [25 + (i % 40) for i in range(n)],
            "salary": [50000.0 + i * 100.0 for i in range(n)],
            "department": ["Eng", "Sales", "HR", "Finance"][0:1] * n if n == 0 else ["Eng"] * n,
        }
    )


BASE_URL = "https://beacon.mirage.local/track"


# --- Yardımcı: xlsx zip'inden rels içeriğini çıkar ---------------------------
def read_zip_member(zip_bytes: bytes, member_path: str) -> str:
    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
        with zf.open(member_path) as f:
            return f.read().decode("utf-8")


def list_zip_members(zip_bytes: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
        return zf.namelist()


def find_member_containing(zip_bytes: bytes, needle: str) -> str | None:
    """İçinde needle geçen ilk zip üyesinin yolunu döndür."""
    for name in list_zip_members(zip_bytes):
        try:
            content = read_zip_member(zip_bytes, name)
            if needle in content:
                return name
        except (UnicodeDecodeError, KeyError):
            continue
    return None


# =============================================================================
# Test Sınıfları
# =============================================================================
class TestHoneytokenOutputValidity(unittest.TestCase):
    """Test grubu 1: Üretilen dosya geçerli bir XLSX mi?"""

    def test_returns_bytes(self):
        """inject_honeytoken bytes döndürmeli (disk I/O yok)."""
        df = make_sample_df(10)
        result = inject_honeytoken(df, base_url=BASE_URL)
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 1000, "xlsx çok küçük — bozuk olabilir")

    def test_output_is_valid_zip(self):
        """Çıktı geçerli bir ZIP (xlsx = zip) olmalı."""
        df = make_sample_df(10)
        result = inject_honeytoken(df, base_url=BASE_URL)
        with zipfile.ZipFile(io.BytesIO(result), "r") as zf:
            self.assertFalse(zf.testzip(), "ZIP corrupt")

    def test_output_is_valid_xlsx(self):
        """Çıktı openpyxl ile açılabilmeli (gerçek xlsx)."""
        from openpyxl import load_workbook

        df = make_sample_df(10)
        result = inject_honeytoken(df, base_url=BASE_URL)
        wb = load_workbook(io.BytesIO(result))
        self.assertIn("Sheet", wb.sheetnames)
        ws = wb["Sheet"]
        # Satır sayısı = veri satırları + 1 (header)
        self.assertEqual(ws.max_row, 11)

    def test_pandas_can_read_without_error(self):
        """Pandas dosyayı sorunsuz okumalı — veri bozulmamış."""
        df = make_sample_df(50)
        result = inject_honeytoken(df, base_url=BASE_URL)
        df_back = pd.read_excel(io.BytesIO(result))
        self.assertEqual(len(df_back), 50)
        # Kolonlar korunmuş
        for col in df.columns:
            self.assertIn(col, df_back.columns)
        # Veri değerleri korunmuş (satır satır)
        pd.testing.assert_frame_equal(
            df.reset_index(drop=True),
            df_back.reset_index(drop=True),
            check_dtype=False,
        )


class TestHoneytokenUrlInjection(unittest.TestCase):
    """Test grubu 2: Tracking URL gerçekten xlsx XML'ine gömülmüş mü?"""

    def test_url_appears_in_zip(self):
        """Tracking URL string olarak xlsx içinde bir yerde olmalı."""
        df = make_sample_df(10)
        result = inject_honeytoken(df, base_url=BASE_URL)
        member = find_member_containing(result, BASE_URL)
        self.assertIsNotNone(
            member,
            f"BASE_URL ({BASE_URL}) hiçbir xlsx üyesinde bulunamadı — enjeksiyon başarısız",
        )

    def test_url_is_in_rels_file(self):
        """URL bir .rels dosyasında olmalı (relationship definition)."""
        df = make_sample_df(10)
        result = inject_honeytoken(df, base_url=BASE_URL)
        members = list_zip_members(result)
        rels_members = [m for m in members if m.endswith(".rels")]
        found = False
        for rm in rels_members:
            content = read_zip_member(result, rm)
            if BASE_URL in content:
                found = True
                break
        self.assertTrue(found, "URL hiçbir .rels dosyasında değil")

    def test_url_has_target_mode_external(self):
        """Relationship TargetMode='External' olmalı (yoksa Excel link'i chok eder)."""
        df = make_sample_df(10)
        result = inject_honeytoken(df, base_url=BASE_URL)
        members = list_zip_members(result)
        for rm in members:
            if not rm.endswith(".rels"):
                continue
            content = read_zip_member(result, rm)
            if BASE_URL in content:
                self.assertIn(
                    'TargetMode="External"',
                    content,
                    f"{rm} içinde URL var ama TargetMode=External değil — Excel link'i takip etmez",
                )

    def test_token_is_unique_per_call(self):
        """Her çağrı farklı bir token üretmeli."""
        df = make_sample_df(10)
        r1 = inject_honeytoken(df, base_url=BASE_URL)
        r2 = inject_honeytoken(df, base_url=BASE_URL)
        # URL'leri çıkar
        url1 = _extract_token_url(r1, BASE_URL)
        url2 = _extract_token_url(r2, BASE_URL)
        self.assertNotEqual(url1, url2, "Aynı token iki kez üretildi — unique değil")

    def test_token_is_uuid_format(self):
        """Token UUID formatında olmalı (kolay parse, çakışma riski sıfır)."""
        df = make_sample_df(10)
        result = inject_honeytoken(df, base_url=BASE_URL)
        url = _extract_token_url(result, BASE_URL)
        # URL'den token kısmını çek
        token = url.rstrip("/").rsplit("/", 1)[-1]
        # UUID regex
        uuid_re = re.compile(
            r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$"
        )
        self.assertIsNotNone(
            uuid_re.match(token),
            f"Token UUID formatında değil: {token!r}",
        )
        # uuid.UUID ile de parse edilebilmeli
        uuid.UUID(token)


def _extract_token_url(zip_bytes: bytes, base_url: str) -> str:
    """Zip'teki rels dosyalarından tracking URL'in tam halini çıkar."""
    members = list_zip_members(zip_bytes)
    for rm in members:
        if not rm.endswith(".rels"):
            continue
        content = read_zip_member(zip_bytes, rm)
        if base_url in content:
            # Target="..." attribute'unu çek
            m = re.search(r'Target="([^"]+)"', content)
            if m:
                return m.group(1)
    raise AssertionError(f"URL bulunamadı: {base_url}")


# =============================================================================
# GÜVENLİK İNVARIANTLARI — en kritik test grubu
# =============================================================================
class TestSecurityInvariants(unittest.TestCase):
    """
    MIRAGE'ın yasal çizgisi: kurban makinede KOD ÇALIŞTIRAN hiçbir şey olamaz.
    Bu testler fail ederse, ürün yasal olarak kullanılamaz.
    """

    def setUp(self):
        df = make_sample_df(20)
        self.result = inject_honeytoken(df, base_url=BASE_URL)
        # Tüm zip üyelerinin içeriğini topla
        self.all_content = []
        for name in list_zip_members(self.result):
            try:
                self.all_content.append((name, read_zip_member(self.result, name)))
            except UnicodeDecodeError:
                pass  # binary dosyalar (PNG vb.) atla

    def test_no_powershell_payload(self):
        for name, content in self.all_content:
            self.assertNotIn(
                "powershell",
                content.lower(),
                f"{name}: PowerShell referansı var — FORBIDDEN",
            )

    def test_no_dde_command_injection(self):
        for name, content in self.all_content:
            self.assertNotIn(
                "cmd|",
                content,
                f"{name}: DDE cmd| injection tespit edildi — FORBIDDEN",
            )

    def test_no_macro_or_vba(self):
        """vbaProject.bin veya macro kodu olmamalı."""
        members = list_zip_members(self.result)
        for m in members:
            self.assertFalse(
                "vbaProject" in m.lower(),
                f"VBA macro projesi bulundu: {m}",
            )
        # İçerikte Auto_Open / Sub / Function gibi VBA ipuçları olmamalı
        vba_patterns = ["Auto_Open", "AutoOpen", "Workbook_Open", "Sub ", "End Sub"]
        for name, content in self.all_content:
            for pat in vba_patterns:
                if pat in content:
                    # XML ad alanlarında "Sub" geçebilir, sadece gerçek VBA pattern'leri için hata ver
                    if pat == "Sub " and "vba" not in name.lower():
                        continue
                    self.fail(f"{name}: VBA pattern '{pat}' bulundu — FORBIDDEN")

    def test_no_dns_tunneling_indicators(self):
        """DNS tunneling veya C2 endpoint paternleri olmamalı."""
        forbidden = ["dns_tunnel", "nslookup", "iodine", "dnscat"]
        for name, content in self.all_content:
            for pat in forbidden:
                self.assertNotIn(
                    pat, content.lower(),
                    f"{name}: DNS tunneling indikatorü '{pat}' — FORBIDDEN",
                )

    def test_no_executable_or_script_extensions_in_zip(self):
        """Zip içinde .exe, .ps1, .vbs, .js, .bat, .cmd olmamalı."""
        forbidden_ext = (".exe", ".ps1", ".vbs", ".js", ".bat", ".cmd", ".scr", ".wsf")
        members = list_zip_members(self.result)
        for m in members:
            for ext in forbidden_ext:
                self.assertFalse(
                    m.lower().endswith(ext),
                    f"Yasak dosya uzantısı tespit edildi: {m}",
                )

    def test_forbidden_patterns_constant_exists(self):
        """Modül, yasaklı pattern'leri export eden bir sabit taşımalı."""
        self.assertIsInstance(MIRAGE_FORBIDDEN_PATTERNS, (list, tuple, set))
        # En azından PowerShell ve DDE orada olmalı
        all_patterns = " ".join(MIRAGE_FORBIDDEN_PATTERNS).lower()
        self.assertIn("powershell", all_patterns)
        self.assertIn("cmd|", all_patterns)

    def test_only_http_or_https_urls(self):
        """Gömülen URL'ler sadece http(s) olmalı (file://, javascript:, smb:// yasak)."""
        url_pattern = re.compile(
            r'Target="([^"]*)"'
        )
        for name, content in self.all_content:
            for m in url_pattern.finditer(content):
                target = m.group(1)
                if "mirage" not in target and "beacon" not in target:
                    # Diğer relationship'ler (iç resimler vb.) — atla
                    continue
                self.assertTrue(
                    target.startswith("http://") or target.startswith("https://"),
                    f"{name}: URL şeması yasaklı: {target}",
                )


# =============================================================================
# Registry testleri — Task 03 ile köprü
# =============================================================================
class TestHoneytokenRegistry(unittest.TestCase):
    """Token'lar kayıt altına alınmalı (beacon geldiğinde eşleştirme için)."""

    def test_registry_records_token(self):
        reg = HoneytokenRegistry()
        df = make_sample_df(10)
        record = reg.issue(df, base_url=BASE_URL)
        self.assertIsInstance(record, HoneytokenRecord)
        self.assertIsNotNone(record.token)
        self.assertEqual(record.base_url, BASE_URL)
        # Token registry'de bulunmalı
        lookup = reg.lookup(record.token)
        self.assertIsNotNone(lookup)
        self.assertEqual(lookup.token, record.token)

    def test_registry_returns_none_for_unknown_token(self):
        reg = HoneytokenRegistry()
        self.assertIsNone(reg.lookup("nonexistent-token"))

    def test_registry_records_metadata(self):
        """Token ile birlikte kaynak DataFrame'in özeti saklanmalı."""
        reg = HoneytokenRegistry()
        df = make_sample_df(25)
        record = reg.issue(df, base_url=BASE_URL, label="Q4-finance-export")
        self.assertEqual(record.label, "Q4-finance-export")
        self.assertEqual(record.row_count, 25)
        self.assertEqual(record.columns, list(df.columns))
        # lookup aynı bilgileri döndürmeli
        lookup = reg.lookup(record.token)
        self.assertEqual(lookup.label, "Q4-finance-export")
        self.assertEqual(lookup.row_count, 25)

    def test_registry_persistence_to_disk(self):
        """Registry diske kaydedilebilmeli (servis restart'ından sonra korunmalı)."""
        import tempfile, os, json
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = f.name
        try:
            reg1 = HoneytokenRegistry(path=path)
            df = make_sample_df(10)
            rec = reg1.issue(df, base_url=BASE_URL, label="test-export")
            reg1.save()
            # Yeni registry örneği diski okusun
            reg2 = HoneytokenRegistry(path=path)
            reg2.load()
            lookup = reg2.lookup(rec.token)
            self.assertIsNotNone(lookup, "Disk persistence başarısız")
            self.assertEqual(lookup.label, "test-export")
        finally:
            os.unlink(path)


class TestIntegrationWithTask01(unittest.TestCase):
    """Task 01 (sentetik veri) → Task 02 (honeytoken) entegrasyonu."""

    def test_synthetic_data_can_be_honeytokenized(self):
        """Sentetik veri üret → honeytoken enjekte → geçerli xlsx al."""
        from mirage import MirageSynthesizer

        # Küçük gerçek veri örneği
        real = pd.DataFrame(
            {
                "user_id": [f"u{i}" for i in range(20)],
                "age": [25 + i for i in range(20)],
                "salary": [50000.0 + i * 1000 for i in range(20)],
                "dept": ["Eng", "Sales"] * 10,
            }
        )
        engine = MirageSynthesizer(seed=7).fit(real)
        synthetic = engine.synthesize(100).df
        # Şimdi honeytoken enjekte et
        result = inject_honeytoken(synthetic, base_url=BASE_URL)
        # Geçerli xlsx ve veri korunmuş
        df_back = pd.read_excel(io.BytesIO(result))
        self.assertEqual(len(df_back), 100)
        # URL injection da çalışıyor
        self.assertIsNotNone(find_member_containing(result, BASE_URL))


# =============================================================================
# Tek dosya çalıştırma desteği (pytest yoksa)
# =============================================================================
if __name__ == "__main__":
    unittest.main(verbosity=2)
